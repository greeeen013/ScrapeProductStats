import requests
from bs4 import BeautifulSoup
import time
from urllib.parse import urljoin
import openpyxl
from openpyxl import Workbook
import os


def ziskej_vyrobce():
    """Získá seznam všech výrobců z hlavní stránky"""
    url = "https://www.myprojectorlamps.eu"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    select = soup.find('select', {'id': 'brands-select'})
    options = select.find_all('option')[1:]  # První option je "Brand", přeskočíme

    vyrobci = []
    for option in options:
        vyrobci.append({
            'value': option['value'],
            'nazev': option.text.strip()
        })
    return vyrobci


def ziskej_produkty_vyrobce(vyrobce_nazev):
    """Získá všechny produkty konkrétního výrobce"""
    url = f"https://www.myprojectorlamps.eu/projectors/{vyrobce_nazev}"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    select = soup.find('select', {'id': 'lamps-select'})
    if not select:
        return []

    produkty = []
    for option in select.find_all('option')[1:]:  # Přeskočíme první option
        if option['value'] != '-':
            produkty.append(option['value'])

    return produkty


def zpracuj_produkt(url, ws, row_index):
    """Zpracuje detail produktu a zapíše data do Excelu"""
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Získání technických informací
    tech_info = soup.find('div', class_='product-table-tech-info')
    if not tech_info:
        return row_index, False

    # Kontrola Lamp Part Number
    part_number = None
    brand = None
    for row in tech_info.find_all('tr'):
        cells = row.find_all('td')
        if len(cells) >= 2:
            if cells[0].text.strip() == 'Lamp Part Number':
                part_number = cells[1].text.strip()
            elif cells[0].text.strip() == 'Brand':
                brand = cells[1].text.strip()

    if part_number == 'See "Alternative Lamp ID\'s"':
        return row_index, False

    # Získání kompatibilních projektorů
    projektory_section = soup.find('div', class_='suitable-projectors-minimalistic')
    kompatibilni = []
    if projektory_section:
        for li in projektory_section.find_all('li'):
            kompatibilni.append(li.text.strip())

    # Zápis do Excelu
    ws[f'A{row_index}'] = brand
    ws[f'B{row_index}'] = part_number
    ws[f'C{row_index}'] = '; '.join(kompatibilni)

    return row_index + 1, True


def nacti_nebo_vytvor_excel(soubor):
    """Načte existující Excel soubor nebo vytvoří nový"""
    if os.path.exists(soubor):
        wb = openpyxl.load_workbook(soubor)
        ws = wb.active
        # Najdeme první volný řádek
        row_index = 1
        while ws[f'A{row_index}'].value is not None:
            row_index += 1
    else:
        wb = Workbook()
        ws = wb.active
        # Přidáme hlavičku
        ws['A1'] = 'Brand'
        ws['B1'] = 'Lamp Part Number'
        ws['C1'] = 'Suitable Projectors'
        row_index = 2

    return wb, ws, row_index


def main():
    # Inicializace Excel souboru
    soubor = 'vysledky.xlsx'
    wb, ws, row_index = nacti_nebo_vytvor_excel(soubor)

    # Získání výrobců
    vyrobci = ziskej_vyrobce()

    # Výběr výrobce
    print("Dostupné značky:")
    for i, vyrobce in enumerate(vyrobci, 1):
        print(f"{i}. {vyrobce['nazev']}")

    volba = input("Zadejte číslo výrobce, název výrobce nebo 'vše' pro všechny: ")

    if volba.lower() == 'vše':
        vybrani_vyrobci = vyrobci
    else:
        # Zkusíme najít podle názvu
        najity_vyrobce = None
        for vyrobce in vyrobci:
            if vyrobce['nazev'].lower() == volba.lower():
                najity_vyrobce = vyrobce
                break

        # Pokud nenajdeme podle názvu, zkusíme podle čísla
        if not najity_vyrobce:
            try:
                index = int(volba) - 1
                if 0 <= index < len(vyrobci):
                    najity_vyrobce = vyrobci[index]
            except ValueError:
                pass

        if najity_vyrobce:
            vybrani_vyrobci = [najity_vyrobce]
            print(f"Vybraný výrobce: {najity_vyrobce['nazev']}")
        else:
            print("Neznámý výrobce!")
            return

    # Zpracování vybraných výrobců
    for vyrobce in vybrani_vyrobci:
        print(f"Zpracovávám {vyrobce['nazev']}...")

        produkty = ziskej_produkty_vyrobce(vyrobce['nazev'])
        print(f"Našel jsem {len(produkty)} produktů")

        for produkt_url in produkty:
            print(f"  Zpracovávám {produkt_url}")
            row_index, uspech = zpracuj_produkt(produkt_url, ws, row_index)
            if uspech:
                # Průběžně ukládáme
                wb.save(soubor)
                time.sleep(1)  # Ohleduplné intervaly mezi requesty

    # Finální uložení
    wb.save(soubor)
    print(f"Data uložena do {soubor}")


if __name__ == "__main__":
    main()